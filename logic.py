

# @title Nombre del archivo
# nombre_del_archivo = 'SOLICITUDES 17 DE JUNIO.xlsx' #@param {type:"string"}

# @title Autenticación
from google.colab import drive
drive.mount('/content/drive')
!pip install --upgrade openpyxl

# @title Procesar archivo completo
# Importando dependencias
import pandas as pd
import shutil
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import os

# Parámetros
#nombre_del_archivo = 'SOLICITUDES 17 DE JUNIO.xlsx'
file_path = "/content/drive/MyDrive/Domicilios/Solicitudes diarias/" + nombre_del_archivo

# Verificar existencia del archivo
if not os.path.exists(file_path):
    raise FileNotFoundError(f"File not found: {file_path}. Please check the file name and path.")

# Leer Excel
data = pd.read_excel(file_path, engine='openpyxl')

# Lista para almacenar los datos procesados
resultados = []
usuarios_dict = {}

# Prefijos que buscamos en Observación
PREFIJOS = {
    'acervo':        'Acervo:',
    'titulo':        'Título:',
    'clasificacion': 'Classificação:',
    'origem':        'Biblioteca de origem:',
    'recibo':        'Biblioteca de recebimento:',
    'telefono':      'Telefone:',
    'email':         'E-mail:',
    'domicilio':     'Domicílio:',
    'fecha':         'Data de solicitação:'
}

for idx, fila in data.iterrows():


    # 1) Separar Usuario y Observación
    columna1 = fila['Usuario / Solicitante']
    columna2 = fila['Observación']

    # 2) Split y limpieza
    partes = [p.strip() for p in columna2.split("|") if p.strip()]
    print(f"\nFila {idx}: {partes}")

    # 3) Construir dict de campos
    campos = {}
    for p in partes:
        for key, pref in PREFIJOS.items():
            # pref puede ser tuple o str
            opciones = pref if isinstance(pref, (list, tuple)) else (pref,)
            for op in opciones:
                if p.startswith(op):
                    valor = p[len(op):].strip()
                    campos[key] = valor
                    break

    # 4) Validar que tengamos lo esencial
    esenciales = ['titulo', 'clasificacion', 'origem', 'recibo',
                   'telefono', 'email', 'domicilio']

    # 5) Añadir recibo y/o Domicilio si hace falta
    if 'recibo' not in campos:
        campos['recibo'] = ''
    else:
      campos['recibo'] = "Biblioteca pública " + campos['recibo']
    if 'domicilio' not in campos:
        campos['domicilio'] = ''

    if not all(k in campos for k in esenciales):
        print(f"Registro {idx} incompleto (faltan {set(esenciales)-campos.keys()}), se excluye.")
        continue

    print(f"Registro {idx} procesando...")

    # 6) Extraer Nombre y Cédula
    try:
        cedula_raw, nombre_raw = columna1.split("-", 1)
    except ValueError:
        print(f"Usuario sin formato en fila {idx}, salto.")
        continue
    Cedula = cedula_raw.lstrip("0").strip()
    Nombre = nombre_raw.strip()


    # 7) Título y demás campos
    Titulo = campos['titulo'] #campos['titulo'].split("/")[0].strip()
    Telefono  = campos['telefono']
    Correo    = campos['email']
    Direccion = campos['domicilio'] + " " + campos['recibo']
    Fecha     = campos.get('fecha', '')
    Topografico = campos['clasificacion']


    # 8) Agregar a dict por usuario para concatenar títulos si repite
    if Cedula in usuarios_dict:
        usuarios_dict[Cedula]['Topografico'] += "\n " + Topografico + "  " + Titulo
    else:
        usuarios_dict[Cedula] = {
            'Nombre': Nombre,
            'Cedula': Cedula,
            'Direccion': Direccion,
            'Telefono': Telefono,
            'Topografico': Topografico + "  " + Titulo
        }

    # 9) Agregar a la lista de resultados
    resultados.append({
        'Nombre': Nombre,
        'Cedula': Cedula,
        'Direccion': Direccion,
        'Localidad': "",
        'Barrio': "",
        'Telefono': Telefono,
        'Topografico': Topografico + "  " + Titulo
    })
    print(f"Fila {idx} procesada.", campos  )


# 10) Convertir a DataFrame y generar Excel formateado
df = pd.DataFrame(resultados)
wb = Workbook()
ws = wb.active

# Encabezados
for col_idx, header in enumerate(df.columns, 1):
    c = ws.cell(row=1, column=col_idx, value=header)
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="center")

# Datos
for row_idx, row in enumerate(df.itertuples(index=False), 2):
    for col_idx, value in enumerate(row, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Ajuste de ancho
for i, col in enumerate(ws.columns, 1):
    max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
    ws.column_dimensions[chr(64 + i)].width = max_len + 2

# Guardar y copiar
archivo_generado = '/content/' + nombre_del_archivo.replace(".xlsx", "_procesado.xlsx")
carpeta_destino   = '/content/drive/MyDrive/Domicilios/Solicitudes generadas'

wb.save(archivo_generado)
shutil.copy(archivo_generado, carpeta_destino)

# Desmontar unidad y fin
#import drive
drive.flush_and_unmount()
print("Proceso completado. Archivo guardado en:", carpeta_destino)
