#!/usr/bin/env python3
import sys
import os
import pandas as pd
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Prefijos a buscar en la columna Observación
PREFIJOS = {
    'acervo':        'Acervo:',
    'titulo':        'Título:',
    'clasificacion': 'Classificação:',
    'origem':        'Biblioteca de origem:',
    'recibo':        'Biblioteca de recebimento:',
    'telefono':      'Telefone:',
    'email':         'E-mail:',
    'domicilio':     'Domicílio:',
    'fecha':         'Data de solicitação:'
}

# Campos estrictamente requeridos (si faltan, se omite)
REQUIRED = ['titulo', 'clasificacion', 'telefono', 'email']

# Para almacenar registros omitidos
omitted = []

def procesar_archivo(nombre_del_archivo):
    file_path = f"/content/drive/MyDrive/Domicilios/Solicitudes diarias/{nombre_del_archivo}"
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    data = pd.read_excel(file_path, engine='openpyxl')
    resultados    = []
    usuarios_dict = {}

    for idx, fila in data.iterrows():
        usu = fila.get('Usuario / Solicitante', '')
        obs = fila.get('Observación', '')
        partes = [p.strip() for p in str(obs).split("|") if p.strip()]
        campos = {}

        for p in partes:
            for key, pref in PREFIJOS.items():
                if p.startswith(pref):
                    campos[key] = p[len(pref):].strip()

        # Verificar campos requeridos
        missing = [f for f in REQUIRED if f not in campos]
        if missing:
            reason = f"Faltan campos requeridos: {', '.join(missing)}"
            omitted.append({'fila': idx+2, 'usuario': usu, 'razon': reason})
            continue

        # Aceptar si al menos existe domicilio o recibo
        if 'domicilio' not in campos and 'recibo' not in campos:
            reason = "Sin domicilio ni biblioteca de recibo"
            omitted.append({'fila': idx+2, 'usuario': usu, 'razon': reason})
            continue

        # Procesar Usuario
        try:
            cc_raw, nombre_raw = usu.split("-", 1)
        except ValueError:
            omitted.append({'fila': idx+2, 'usuario': usu, 'razon': 'Formato de usuario inválido'})
            continue

        Cedula = cc_raw.lstrip("0").strip()
        Nombre = nombre_raw.strip()

        # Construir Direccion: domicilio + recibo (o solo recibo si falta domicilio)
        direccion_parts = []
        if 'domicilio' in campos:
            direccion_parts.append(campos['domicilio'])
        if 'recibo' in campos:
            direccion_parts.append(f"Recibe en: {campos['recibo']}")
        Direccion = " | ".join(direccion_parts)

        Titulo      = campos['titulo']
        Telefono    = campos['telefono']
        Correo      = campos['email']
        Topografico = campos['clasificacion']

        # Acumular por usuario si es necesario
        if Cedula in usuarios_dict:
            usuarios_dict[Cedula]['Topografico'] += f"\n{Topografico}  {Titulo}"
        else:
            usuarios_dict[Cedula] = {
                'Nombre':      Nombre,
                'Cedula':      Cedula,
                'Direccion':   Direccion,
                'Telefono':    Telefono,
                'Topografico': f"{Topografico}  {Titulo}"
            }

        resultados.append({
            'Nombre':      Nombre,
            'Cedula':      Cedula,
            'Direccion':   Direccion,
            'Localidad':   "",
            'Barrio':      "",
            'Telefono':    Telefono,
            'Topografico': f"{Topografico}  {Titulo}"
        })

    # Mostrar registros omitidos con razón
    if omitted:
        print("\n--- Solicitudes omitidas ---")
        for om in omitted:
            print(f"Fila {om['fila']}: Usuario={om['usuario']} -> {om['razon']}")
        print("--- Fin omitidos ---\n")

    # Generar DataFrame y guardar
    df = pd.DataFrame(resultados)
    wb = Workbook()
    ws = wb.active

    for i, h in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font      = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    for r, row in enumerate(df.itertuples(index=False), 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[chr(64+col_idx)].width = max_len + 2

    salida  = nombre_del_archivo.replace(".xlsx", "_procesado.xlsx")
    destino = "/content/" + salida
    carpeta = "/content/drive/MyDrive/Domicilios/Solicitudes generadas"
    wb.save(destino)
    shutil.copy(destino, carpeta)
    print("✅ Proceso completado. Archivo en:", carpeta)

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Uso: python logic.py <nombre_del_archivo.xlsx>")
        sys.exit(1)
    procesar_archivo(sys.argv[1])
